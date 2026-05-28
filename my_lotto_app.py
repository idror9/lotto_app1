import tkinter as tk
from tkinter import messagebox, filedialog
import win32com.client
import pandas as pd
import os
import sys
from openpyxl.styles import Alignment, Border, Side

# פונקציה שמאפשרת לקוד למצוא את האייקון גם כשהוא בתוך ה-EXE
def resource_path(relative_path):
    try:
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")
    return os.path.join(base_path, relative_path)

def get_desktop_path():
    return os.path.join(os.path.expanduser('~'), 'Desktop')

def force_paste(widget):
    try:
        text = widget.clipboard_get()
        if widget.selection_present():
            widget.delete("sel.first", "sel.last")
        widget.insert(tk.INSERT, text)
    except:
        pass
    return "break"

def force_copy(widget):
    try:
        if widget.selection_present():
            text = widget.get()[widget.index("sel.first"):widget.index("sel.last")]
            widget.clipboard_clear()
            widget.clipboard_append(text)
    except:
        pass
    return "break"

# עדכון התווית של תיבת המחשב/אזור לפי סוג התקלה
def update_label(*args):
    if issue_type.get() == "תקלת מסך":
        computer_label.config(text="שם אזור:")
    else:
        computer_label.config(text="שם מחשב:")

def create_outlook_mail(subject, html_body):
    try:
        outlook = win32com.client.Dispatch("outlook.application")
        mail = outlook.CreateItem(0)
        mail.To = "david_yafit@meuhedet.co.il"
        mail.Subject = subject
        mail.BodyFormat = 2  
        mail.HTMLBody = f'<div style="direction: rtl; text-align: right;">{html_body}</div>'
        mail.Display(True)
    except Exception as e:
        messagebox.showerror("שגיאה", str(e))

def send_email_logic():
    clinic = clinic_name.get()
    address = clinic_address.get()
    contact = contact_person.get()
    call = call_reason.get()
    comp_or_area = computer_name.get()
    issue = issue_type.get()
    
    if issue == "תקלת מחשב":
        body = f"היי יפית, בהמשך לקריאה {call}<br>אבקש לשלוח טכנאי למרפאת {clinic}<br>כתובת {address}<br>מהות התקלה תקלת מחשב שם מחשב {comp_or_area}<br>ציוד נדרש מחשב קיופלו עם נגן מדיה פלייר לא M70<br>איש קשר {contact}"
    elif issue == "תקלת מסך":
        body = f"היי יפית, בהמשך לקריאה {call}<br>אבקש לשלוח טכנאי למרפאת {clinic}<br>כתובת {address}<br>מהות התקלה תקלת מסך {comp_or_area}<br>איש קשר {contact}"
    else:
        body = f"היי יפית, בהמשך לקריאה {call}<br>אבקש לשלוח טכנאי למרפאת {clinic}<br>כתובת {address}<br>מהות התקלה תקלת מקלט משדר שם מחשב {comp_or_area}<br>איש קשר {contact}"
    
    create_outlook_mail(f"שליחת טכנאי למרפאת {clinic}", body)

def send_to_matrix():
    try:
        file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx *.xls")])
        if not file_path: return
        df = pd.read_excel(file_path)
        cols_to_keep = ['מספר קריאת ספק', 'תאריך עדכון אחרון', 'תאור', 'כתובת', 'לקוח', 'טלפון נייד']
        new_df = df[[c for c in cols_to_keep if c in df.columns]].copy()
        new_df = new_df.rename(columns={'תאריך עדכון אחרון': 'תאריך שליחת הקריאה למטריקס'})
        new_df['סטטוס'] = ''
        output_path = os.path.join(get_desktop_path(), 'output_matrix.xlsx')
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            new_df.to_excel(writer, index=False, sheet_name='Matrix')
            workbook = writer.book
            worksheet = workbook.active
            worksheet.sheet_view.rightToLeft = True
            thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
            col_indices = {cell.value: i+1 for i, cell in enumerate(worksheet[1])}
            desc_col = col_indices.get('תאור')
            phone_col = col_indices.get('טלפון נייד')
            if desc_col:
                column_letter = worksheet.cell(row=1, column=desc_col).column_letter
                worksheet.column_dimensions[column_letter].width = 60
            for row_idx, row in enumerate(worksheet.iter_rows(min_row=1, max_row=worksheet.max_row), 1):
                for cell in row:
                    cell.border = thin_border
                    align_params = {'vertical': 'top', 'horizontal': 'right'}
                    if cell.column == desc_col: align_params['wrapText'] = True
                    if cell.column == phone_col and row_idx > 1:
                        cell.number_format = '@' 
                        if cell.value is not None:
                            val = str(cell.value).replace('.0', '').strip()
                            if val and not val.startswith('0') and len(val) >= 9: val = '0' + val
                            cell.value = val
                    cell.alignment = Alignment(**align_params)
        os.startfile(output_path)
    except Exception as e:
        messagebox.showerror("שגיאה", str(e))

# פונקציה לביצוע פינג רציף (ping -t)
def run_ping():
    target = ping_entry.get().strip()
    if not target:
        messagebox.showwarning("אזהרה", "אנא הקלד כתובת IP או שם מחשב לבדיקת פינג")
        return
    # שימוש בגרשיים כפולים מסביב לפקודה ודגל -t לפינג קבוע
    os.system(f'start cmd /k "ping {target} -t"')

# פונקציה להעתקת כתובת קיופלו ללוח
def copy_qflow_url():
    url = "https://qfn.meuhedet.org/qflow"
    root.clipboard_clear()
    root.clipboard_append(url)
    messagebox.showinfo("העתקה", "הכתובת הועתקה בהצלחה ללוח!")

# פונקציית ניקוי מסודרת ובטוחה לכל השדות
def clear_all_fields():
    fields = [clinic_name, clinic_address, computer_name, contact_person, call_reason, ping_entry]
    for field in fields:
        field.delete(0, tk.END)

# הגדרת הממשק הראשי
root = tk.Tk()
root.title("מערכת תקלות קיופלו - 3.13")
root.geometry("450x600") 
root.configure(bg='#f0f0f0')

# הגדרת האייקון לחלון
try:
    icon_path = resource_path("qflow.ico")
    root.iconbitmap(icon_path)
except Exception as e:
    print(f"לא ניתן היה לטעון את האייקון: {e}")

root.columnconfigure(0, weight=1)
root.columnconfigure(1, weight=1)

def setup_entry(row, label_text):
    lbl = tk.Label(root, text=label_text, bg='#f0f0f0')
    lbl.grid(row=row, column=1, padx=10, pady=5, sticky='e')
    en = tk.Entry(root, justify='right')
    en.grid(row=row, column=0, padx=10, pady=5, sticky='ew')
    en.bind("<Button-3>", lambda e: force_paste(en))
    en.bind("<F12>", lambda e: force_paste(en))
    en.bind("<Control-v>", lambda e: force_paste(en))
    en.bind("<Control-V>", lambda e: force_paste(en))
    en.bind("<Control-c>", lambda e: force_copy(en))
    en.bind("<Control-a>", lambda e: [en.select_range(0, 'end'), "break"][1])
    return en, lbl

# שורת בחירת סוג תקלה
tk.Label(root, text="בחר סוג תקלה:", bg='#f0f0f0').grid(row=0, column=1, padx=10, pady=15, sticky='e')
issue_type = tk.StringVar(value="תקלת מחשב")
issue_type.trace("w", update_label)
tk.OptionMenu(root, issue_type, "תקלת מחשב", "תקלת מסך", "תקלת מקלט משדר").grid(row=0, column=0, padx=10, pady=15, sticky='ew')

# שדות הקלט הסטנדרטיים
clinic_name = setup_entry(1, "שם מרפאה:")[0]
clinic_address = setup_entry(2, "כתובת:")[0]
computer_name, computer_label = setup_entry(3, "שם מחשב:")
contact_person = setup_entry(4, "איש קשר:")[0]
call_reason = setup_entry(5, "קריאה:")[0]

# --- שורת פינג ---
tk.Label(root, text="כתובת לפינג:", bg='#f0f0f0').grid(row=6, column=1, padx=10, pady=10, sticky='e')
ping_frame = tk.Frame(root, bg='#f0f0f0')
ping_frame.grid(row=6, column=0, padx=10, pady=10, sticky='ew')

ping_entry = tk.Entry(ping_frame, justify='right', width=15)
ping_entry.pack(side='right', fill='x', expand=True, padx=(0, 5))
ping_entry.bind("<Button-3>", lambda e: force_paste(ping_entry))
ping_entry.bind("<Control-v>", lambda e: force_paste(ping_entry))
ping_entry.bind("<Control-V>", lambda e: force_paste(ping_entry))

ping_btn = tk.Button(ping_frame, text="שלח פינג", command=run_ping, bg="#9b59b6", fg="white", font=("Arial", 9, "bold"))
ping_btn.pack(side='left')

# --- שורת כפתורי פעולה 1 (מטריקס והעתקת כתובת) ---
tk.Button(root, text="SEND2MATRIX", command=send_to_matrix, bg="#3498db", fg="white", width=18).grid(row=7, column=1, padx=10, pady=10, sticky='e')
tk.Button(root, text="העתק כתובת QFlow", command=copy_qflow_url, bg="#f39c12", fg="white", width=18).grid(row=8, column=1, padx=10, pady=5, sticky='e')

# --- שורת כפתורי פעולה 2 (אאוטלוק) ---
tk.Button(root, text="שלח מייל (Outlook)", command=send_email_logic, bg="#2ecc71", fg="white", width=18).grid(row=7, column=0, rowspan=2, padx=10, pady=10, sticky='w')

# --- כפתור ניקוי תחתון ---
tk.Button(root, text="נקה הכל", command=clear_all_fields, bg="#e74c3c", fg="white", width=22).grid(row=9, column=0, columnspan=2, pady=20)

root.mainloop()
